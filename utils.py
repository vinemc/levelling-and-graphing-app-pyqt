import logging
import csv
from pathlib import Path
import json
import datetime
import tempfile
import openpyxl
from fpdf import FPDF

# --- CONSTANTS ---
DEFAULT_ROW_COUNT = 30
SCROLL_ROW_ADD = 10
MAX_SANE_READING = 5.0
SMOOTH_CURVE_POINTS = 300
STATUS_BAR_CLEAR_DELAY = 4000
INPUT_VALIDATION_HIGHLIGHT_DELAY = 3000
APP_VERSION = "1.0.0"

# --- Utility Functions ---
def is_number(val: str) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False

def format_num(val, precision=3):
    try:
        return f"{float(val):.{precision}f}"
    except (ValueError, TypeError):
        return val

# --- Tooltip class for UI help ---
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert") if hasattr(self.widget, "bbox") else (0,0,0,0)
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = ttk.Label(tw, text=self.text, justify=tk.LEFT,
                          background=getattr(Tooltip, 'bg', "#ffffe0"),
                          foreground=getattr(Tooltip, 'fg', "#222"),
                          relief=tk.SOLID, borderwidth=1,
                          font=("tahoma", 8, "normal"))
        label.pack(ipadx=1)

    def hide(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


def generate_pdf_report(results_data, fig, file_path):
    """Generates a comprehensive PDF report of the leveling data and graph."""
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Leveling Survey Report", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 5, f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(10)
        
        if results_data:
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 10, "Calculation Results", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "B", 9)
            
            header = list(results_data[0].keys())
            col_widths = [pdf.w / (len(header) + 1)] * len(header)

            for i, header_text in enumerate(header):
                pdf.cell(col_widths[i], 7, header_text, 1, new_x="RIGHT", new_y="TOP", align="C")
            pdf.ln()

            pdf.set_font("Helvetica", "", 9)
            for row in results_data:
                for i, key in enumerate(header):
                    pdf.cell(col_widths[i], 6, str(row.get(key, "")), 1, new_x="RIGHT", new_y="TOP", align="C")
                pdf.ln()
            pdf.ln(10)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Profile Graph", new_x="LMARGIN", new_y="NEXT")
        
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_img:
            graph_image_path = tmp_img.name
            fig.savefig(graph_image_path, dpi=300)
        
        try:
            pdf.image(graph_image_path, x=10, y=pdf.get_y(), w=190)
        finally:
            import os
            if os.path.exists(graph_image_path):
                try:
                    os.remove(graph_image_path)
                except Exception as e:
                    logging.warning(f"Could not remove temp file: {e}")
        
        pdf.output(file_path)
        return True
    except Exception as e:
        logging.error(f"Failed to generate PDF report: {e}")
        raise

def export_to_excel(data, file_path):
    """Export data to Excel (.xlsx) file."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Could not create worksheet")
        
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        
        wb.save(file_path)
        return True
    except Exception as e:
        logging.error(f"Failed to export to Excel: {e}")
        raise

def save_session(data, settings):
    """Save current session (data, settings, etc.)"""
    session = {
        "data": data,
        "settings": settings.copy(),
    }
    with Path("last_session.json").open("w") as f:
        json.dump(session, f)

def load_session():
    """Load session data from file"""
    try:
        with Path("last_session.json").open("r") as f:
            session = json.load(f)
        return session
    except Exception as e:
        logging.error(f"Could not load session: {e}")
        return None

def update_recent_files(file_path, settings):
    """Add a file to the recent files list and update settings."""
    if not file_path:
        return
    file_path = str(file_path)
    recent = settings.get("recent_files", [])
    if file_path in recent:
        recent.remove(file_path)
    recent.insert(0, file_path)
    settings["recent_files"] = recent[:10]
    return settings
