from PyQt6.QtWidgets import QFileDialog, QMessageBox, QDialog, QProgressBar, QVBoxLayout, QRadioButton, QButtonGroup, QLabel, QPushButton, QWidget, QTableWidgetItem, QLineEdit, QTextEdit, QHBoxLayout, QFileDialog as QtFileDialog, QDialogButtonBox, QCheckBox, QComboBox, QGroupBox, QFormLayout
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QCloseEvent
from pathlib import Path
import csv
import openpyxl
from .utils_qt import ImportDialog
from .lang import LANG
import datetime
import tempfile
from fpdf import FPDF
from PyQt6.QtGui import QPixmap
import qrcode
from PyQt6.QtGui import QImage
import PyPDF2
import subprocess
import sys
import logging
import os
from PyQt6.QtGui import QIcon

# --- Configure logging ---
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
# Create a file handler
handler = logging.FileHandler('app.log', mode='a')
handler.setLevel(logging.DEBUG)
# Create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
# Add the handlers to the logger
if not logger.handlers:
    logger.addHandler(handler)


# --- Fix PyPDF2 import for linter and runtime robustness ---
try:
    import PyPDF2  # type: ignore[import]
except ImportError:
    PyPDF2 = None
    print("PyPDF2 is not installed. PDF password and metadata features will be disabled.")

class ImportExportManager:
    def __init__(self, master, settings, save_settings_callback):
        self.master = master
        self.settings = settings
        self.save_settings_callback = save_settings_callback

    def import_leveling_csv(self, table_widget, column_names, redraw_callback, progress_bar, file_path=None):
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(self.master, "Open Leveling CSV", "", "CSV Files (*.csv)")
        if not file_path:
            return
        self._update_recent_files(file_path)
        dialog = ImportDialog(self.master, file_path, column_names)
        if not dialog.exec():
            return
        if dialog.import_result is None:
            return
        def handler(all_rows, mapping):
            try:
                table_widget.blockSignals(True)
                table_widget.setRowCount(0)
                for row_idx, row_data in enumerate(all_rows):
                    table_widget.insertRow(row_idx)
                    for col_name, col_idx in mapping.items():
                        if col_idx is not None and col_idx < len(row_data):
                            target_idx = column_names.index(col_name)
                            table_widget.setItem(row_idx, target_idx, QTableWidgetItem(row_data[col_idx]))
                if redraw_callback:
                    redraw_callback()
            except Exception as e:
                QMessageBox.critical(self.master, "Import Error", f"Failed to import CSV data:\n{e}")
            finally:
                table_widget.blockSignals(False)
        self._perform_csv_import(file_path, dialog.import_result["mapping"], dialog.import_result["has_header"], handler, progress_bar)

    def import_profile_csv(self, table_widget, column_names, redraw_callback, progress_bar, file_path=None):
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(self.master, "Open Profile CSV", "", "CSV Files (*.csv)")
        if not file_path:
            return
        self._update_recent_files(file_path)
        dialog = ImportDialog(self.master, file_path, column_names)
        if not dialog.exec():
            return
        if dialog.import_result is None:
            return
        def handler(all_rows, mapping):
            try:
                table_widget.blockSignals(True)
                table_widget.setRowCount(0)
                for row_idx, row_data in enumerate(all_rows):
                    table_widget.insertRow(row_idx)
                    for col_name, col_idx in mapping.items():
                        if col_idx is not None and col_idx < len(row_data):
                            target_idx = column_names.index(col_name)
                            table_widget.setItem(row_idx, target_idx, QTableWidgetItem(row_data[col_idx]))
                if redraw_callback:
                    redraw_callback()
            except Exception as e:
                QMessageBox.critical(self.master, "Import Error", f"Failed to import CSV data:\n{e}")
            finally:
                table_widget.blockSignals(False)
        self._perform_csv_import(file_path, dialog.import_result["mapping"], dialog.import_result["has_header"], handler, progress_bar)

    def export_leveling_csv(self, result_table, progress_bar):
        file_path, _ = QFileDialog.getSaveFileName(self.master, "Export Leveling Results", "", "CSV Files (*.csv)")
        if not file_path:
            return
        if progress_bar:
            progress_bar.setVisible(True)
            progress_bar.setRange(0, 0)
        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                headers = [result_table.horizontalHeaderItem(i).text() for i in range(result_table.columnCount())]
                writer.writerow(headers)
                for row in range(result_table.rowCount()):
                    writer.writerow([result_table.item(row, col).text() if result_table.item(row, col) else '' for col in range(result_table.columnCount())])
            self._update_recent_files(file_path)
            QMessageBox.information(self.master, LANG["export_success"], LANG["exported_csv"])
        except Exception as e:
            QMessageBox.critical(self.master, LANG["export_error"], f"{LANG['failed_export_csv']}\n{e}")
        if progress_bar:
            progress_bar.setVisible(False)

    def export_to_excel(self, result_table, graph_table, progress_bar):
        dialog = QDialog(self.master)
        dialog.setWindowTitle(LANG["export_to_excel"])
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(LANG["export_which_table"]))
        results_radio = QRadioButton(LANG["results_table"])
        profile_radio = QRadioButton(LANG["profile_table"])
        results_radio.setChecked(True)
        layout.addWidget(results_radio)
        layout.addWidget(profile_radio)
        btn = QPushButton("Export")
        layout.addWidget(btn)
        btn_group = QButtonGroup(dialog)
        btn_group.addButton(results_radio)
        btn_group.addButton(profile_radio)
        def do_export():
            dialog.accept()
            file_path, _ = QFileDialog.getSaveFileName(self.master, "Export to Excel", "", "Excel Files (*.xlsx)")
            if not file_path:
                return
            if progress_bar:
                progress_bar.setVisible(True)
                progress_bar.setRange(0, 0)
            try:
                table = result_table if results_radio.isChecked() else graph_table
                wb = openpyxl.Workbook()
                ws = wb.active
                if ws is not None:
                    headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    ws.append(headers)
                    for row in range(table.rowCount()):
                        ws.append([table.item(row, col).text() if table.item(row, col) else '' for col in range(table.columnCount())])
                    wb.save(file_path)
                    self._update_recent_files(file_path)
                    QMessageBox.information(self.master, LANG["export_success"], f"{LANG['exported_excel']}\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self.master, LANG["export_error"], f"{LANG['failed_export_excel']}\n{e}")
            if progress_bar:
                progress_bar.setVisible(False)
        btn.clicked.connect(do_export)
        dialog.exec()

    def export_graph(self, fig, file_path=None):
        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(self.master, "Export Graph", "", "PNG Image (*.png);;PDF Document (*.pdf)")
        if not file_path:
            return
        try:
            fig.savefig(file_path)
            QMessageBox.information(self.master, LANG["export_success"], f"{LANG['exported_graph']}\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.master, LANG["export_error"], f"{LANG['failed_export_graph']}\n{e}")

    def export_pdf_with_options(self, result_table, fig):
        from PyQt6.QtWidgets import QMessageBox
        dialog = QDialog(self.master)
        dialog.setWindowTitle("Export to PDF")
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Choose PDF export style:"))
        simple_radio = QRadioButton("Simple PDF (table and graph)")
        professional_radio = QRadioButton("Professional PDF (cover, details, logo, summary, etc.)")
        simple_radio.setChecked(True)
        layout.addWidget(simple_radio)
        layout.addWidget(professional_radio)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
        result = {}
        def accept():
            result['style'] = 'professional' if professional_radio.isChecked() else 'simple'
            dialog.accept()
        def reject():
            dialog.reject()
        buttons.accepted.connect(accept)
        buttons.rejected.connect(reject)
        dialog.exec()
        if dialog.result() != QDialog.DialogCode.Accepted:
            return
        if result['style'] == 'professional':
            options = self.professional_pdf_export_dialog(result_table, fig)
            if not options:
                QMessageBox.warning(self.master, "Export Cancelled", "No options selected for professional export.")
                return
            if options.get('preview'):
                import tempfile, os
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
                    tmp_path = tmp_pdf.name
                self.generate_pdf_report(result_table, fig, professional=True, file_path=tmp_path, **options)
                # Try to open the PDF with the system viewer
                preview_opened = False
                try:
                    if sys.platform.startswith('win'):
                        os.startfile(tmp_path)
                        preview_opened = True
                    elif sys.platform.startswith('darwin'):
                        subprocess.run(['open', tmp_path])
                        preview_opened = True
                    else:
                        subprocess.run(['xdg-open', tmp_path])
                        preview_opened = True
                except Exception:
                    QMessageBox.warning(self.master, "Preview Error", "Could not open PDF preview.")
                if preview_opened:
                    QMessageBox.information(self.master, "Preview Opened", f"Preview PDF opened at: {tmp_path}")
                # Ask user to confirm export and pick location
                reply = QMessageBox.question(self.master, "Export PDF", "Are you satisfied with the preview? Save the final PDF?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    file_path, _ = QFileDialog.getSaveFileName(self.master, "Save PDF Report", "", "PDF Documents (*.pdf)")
                    if file_path:
                        self.generate_pdf_report(result_table, fig, professional=True, file_path=file_path, **options)
                        QMessageBox.information(self.master, "PDF Saved", f"PDF report saved at: {file_path}")
                    else:
                        QMessageBox.information(self.master, "Export Cancelled", "No file selected. PDF not saved.")
                # Clean up temp file
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
            else:
                file_path, _ = QFileDialog.getSaveFileName(self.master, "Save PDF Report", "", "PDF Documents (*.pdf)")
                if file_path:
                    self.generate_pdf_report(result_table, fig, professional=True, file_path=file_path, **options)
                    QMessageBox.information(self.master, "PDF Saved", f"PDF report saved at: {file_path}")
                else:
                    QMessageBox.information(self.master, "Export Cancelled", "No file selected. PDF not saved.")
        else:
            file_path, _ = QFileDialog.getSaveFileName(self.master, "Save PDF Report", "", "PDF Documents (*.pdf)")
            if file_path:
                self.generate_pdf_report(result_table, fig, file_path=file_path)
                QMessageBox.information(self.master, "PDF Saved", f"PDF report saved at: {file_path}")
            else:
                QMessageBox.information(self.master, "Export Cancelled", "No file selected. PDF not saved.")

    def professional_pdf_export_dialog(self, result_table, fig):
        from PyQt6.QtWidgets import QScrollArea
        defaults = self.settings.get('pdf_export_defaults', {})
        dialog = QDialog(self.master)
        dialog.setWindowTitle("Professional PDF Export")
        # Create a scroll area
        scroll = QScrollArea(dialog)
        scroll.setWidgetResizable(True)
        content = QWidget()
        scroll.setWidget(content)
        layout = QVBoxLayout(content)
        # Project details
        details_group = QGroupBox("Project Details")
        details_layout = QFormLayout(details_group)
        project_name_edit = QLineEdit(defaults.get('project_name', ''))
        surveyor_edit = QLineEdit(defaults.get('surveyor', ''))
        date_edit = QLineEdit(defaults.get('date', datetime.datetime.now().strftime('%Y-%m-%d')))
        contact_edit = QLineEdit(defaults.get('contact', ''))
        summary_edit = QTextEdit()
        summary_edit.setPlainText(defaults.get('summary', ''))
        details_layout.addRow("Project Name:", project_name_edit)
        details_layout.addRow("Surveyor:", surveyor_edit)
        details_layout.addRow("Date:", date_edit)
        details_layout.addRow("Contact Info:", contact_edit)
        details_layout.addRow("Executive Summary:", summary_edit)
        # Logo and QR code
        logo_path_edit = QLineEdit(defaults.get('logo_path', ''))
        logo_btn = QPushButton(QIcon('leveling_app_modular/icons/open.svg'), "Browse Logo...")
        def browse_logo():
            path, _ = QtFileDialog.getOpenFileName(dialog, "Select Logo", "", "Image Files (*.png *.jpg *.jpeg *.bmp)")
            if path:
                logo_path_edit.setText(path)
        logo_btn.clicked.connect(browse_logo)
        qr_data_edit = QLineEdit(defaults.get('qr_data', ''))
        details_layout.addRow("Logo:", logo_path_edit)
        details_layout.addRow("", logo_btn)
        details_layout.addRow("QR Code Data (URL):", qr_data_edit)
        layout.addWidget(details_group)
        # Section selection
        section_group = QGroupBox("Sections to Include")
        section_layout = QVBoxLayout(section_group)
        toc_cb = QCheckBox("Table of Contents")
        toc_cb.setChecked(defaults.get('sections', {}).get('toc', True))
        calc_details_cb = QCheckBox("Calculation Details")
        calc_details_cb.setChecked(defaults.get('sections', {}).get('calc_details', False))
        raw_data_cb = QCheckBox("Raw Data Appendix")
        raw_data_cb.setChecked(defaults.get('sections', {}).get('raw_data', False))
        change_log_cb = QCheckBox("Change Log")
        change_log_cb.setChecked(defaults.get('sections', {}).get('change_log', False))
        appendix_cb = QCheckBox("Additional Appendix")
        appendix_cb.setChecked(defaults.get('sections', {}).get('appendix', False))
        section_layout.addWidget(toc_cb)
        section_layout.addWidget(calc_details_cb)
        section_layout.addWidget(raw_data_cb)
        section_layout.addWidget(change_log_cb)
        section_layout.addWidget(appendix_cb)
        layout.addWidget(section_group)
        # Branding and appearance
        branding_group = QGroupBox("Branding & Appearance")
        branding_layout = QFormLayout(branding_group)
        color_scheme_cb = QComboBox()
        color_scheme_cb.addItems(["Default", "Dark", "Light", "Custom..."])
        color_scheme_cb.setCurrentText(defaults.get('color_scheme', 'Default'))
        font_cb = QComboBox()
        font_cb.addItems(["Helvetica", "Arial", "Times", "Courier", "Custom..."])
        font_cb.setCurrentText(defaults.get('font', 'Helvetica'))
        watermark_edit = QLineEdit(defaults.get('watermark', ''))
        bg_img_edit = QLineEdit(defaults.get('bg_img', ''))
        bg_img_btn = QPushButton(QIcon('leveling_app_modular/icons/open.svg'), "Browse Background...")
        def browse_bg():
            path, _ = QtFileDialog.getOpenFileName(dialog, "Select Background Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp)")
            if path:
                bg_img_edit.setText(path)
        bg_img_btn.clicked.connect(browse_bg)
        branding_layout.addRow("Color Scheme:", color_scheme_cb)
        branding_layout.addRow("Font:", font_cb)
        branding_layout.addRow("Watermark Text:", watermark_edit)
        branding_layout.addRow("Background Image:", bg_img_edit)
        branding_layout.addRow("", bg_img_btn)
        layout.addWidget(branding_group)
        # Security and metadata
        security_group = QGroupBox("Security & Metadata")
        security_layout = QFormLayout(security_group)
        password_edit = QLineEdit(defaults.get('password', ''))
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        metadata_edit = QTextEdit()
        metadata_edit.setPlainText(defaults.get('metadata', ''))
        compliance_edit = QTextEdit()
        compliance_edit.setPlainText(defaults.get('compliance', ''))
        security_layout.addRow("PDF Password:", password_edit)
        security_layout.addRow("Custom Metadata:", metadata_edit)
        security_layout.addRow("Compliance/Legal Notices:", compliance_edit)
        layout.addWidget(security_group)
        # Preview and save settings
        preview_cb = QCheckBox("Preview before export")
        preview_cb.setChecked(defaults.get('preview', False))
        save_settings_cb = QCheckBox("Save these settings for future exports")
        save_settings_cb.setChecked(False)
        layout.addWidget(preview_cb)
        layout.addWidget(save_settings_cb)
        # Dialog buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)
        if ok_button:
            ok_button.setIcon(QIcon('leveling_app_modular/icons/export.svg'))
        cancel_button = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        if cancel_button and os.path.exists('leveling_app_modular/icons/exit.svg'):
            cancel_button.setIcon(QIcon('leveling_app_modular/icons/exit.svg'))
        layout.addWidget(buttons)
        result = {}
        def accept():
            result['project_name'] = project_name_edit.text()
            result['surveyor'] = surveyor_edit.text()
            result['date'] = date_edit.text()
            result['contact'] = contact_edit.text()
            result['summary'] = summary_edit.toPlainText()
            result['logo_path'] = logo_path_edit.text()
            result['qr_data'] = qr_data_edit.text()
            result['sections'] = {
                'toc': toc_cb.isChecked(),
                'calc_details': calc_details_cb.isChecked(),
                'raw_data': raw_data_cb.isChecked(),
                'change_log': change_log_cb.isChecked(),
                'appendix': appendix_cb.isChecked(),
            }
            result['color_scheme'] = color_scheme_cb.currentText()
            result['font'] = font_cb.currentText()
            result['watermark'] = watermark_edit.text()
            result['bg_img'] = bg_img_edit.text()
            result['password'] = password_edit.text()
            result['metadata'] = metadata_edit.toPlainText()
            result['compliance'] = compliance_edit.toPlainText()
            result['preview'] = preview_cb.isChecked()
            result['save_settings'] = save_settings_cb.isChecked()
            if save_settings_cb.isChecked():
                self.settings['pdf_export_defaults'] = result.copy()
                self.save_settings_callback()
            dialog.accept()
        def reject():
            dialog.reject()
        buttons.accepted.connect(accept)
        buttons.rejected.connect(reject)
        # Set scroll area as the main layout
        main_layout = QVBoxLayout(dialog)
        main_layout.addWidget(scroll)
        dialog.setMinimumSize(600, 600)
        dialog.exec()
        if dialog.result() != QDialog.DialogCode.Accepted:
            return None
        return result

    def _add_pdf_cover_page(self, pdf, **kwargs):
        pdf.cover_page()

    def _add_pdf_toc(self, pdf, font, sections):
        if sections.get('toc', True):
            pdf.add_page()
            toc_page = pdf.page_no()
            pdf.set_font(font, "B", 16)
            pdf.cell(0, 15, "Table of Contents", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)
            return toc_page
        return None

    def _add_pdf_summary(self, pdf, font, summary, section_pages):
        pdf.add_page()
        section_pages['Executive Summary'] = pdf.page_no()
        if hasattr(pdf, 'bookmark'):
            pdf.bookmark('Executive Summary', level=0)
        pdf.set_font(font, "B", 14)
        pdf.cell(0, 10, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(font, "", 12)
        pdf.multi_cell(0, 8, summary or "No summary provided.")
        pdf.ln(5)

    def _add_pdf_results_table(self, pdf, font, result_table, color_scheme, section_pages):
        pdf.add_page()
        section_pages['Calculation Results'] = pdf.page_no()
        if hasattr(pdf, 'bookmark'):
            pdf.bookmark('Calculation Results', level=0)
        pdf.set_font(font, "B", 14)
        pdf.cell(0, 10, "Calculation Results", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        pdf.set_font(font, "B", 10)
        col_count = result_table.columnCount()
        col_widths = [max(20, int(190 / col_count))] * col_count
        headers = [result_table.horizontalHeaderItem(i).text() if result_table.horizontalHeaderItem(i) else "" for i in range(col_count)]
        for i, header_text in enumerate(headers):
            pdf.cell(col_widths[i], 7, header_text, 1, new_x="RIGHT", new_y="TOP", align="C")
        pdf.ln()
        pdf.set_font(font, "", 10)
        for row in range(result_table.rowCount()):
            if row % 2 == 0:
                pdf.set_fill_color(245, 245, 245) if color_scheme == "Default" else pdf.set_fill_color(40, 40, 40)
            else:
                pdf.set_fill_color(255, 255, 255)
            for i in range(col_count):
                item = result_table.item(row, i)
                value = item.text() if item else ""
                pdf.cell(col_widths[i], 6, str(value), 1, new_x="RIGHT", new_y="TOP", align="C", fill=True)
            pdf.ln()
        pdf.ln(10)

    def _add_pdf_graph(self, pdf, font, fig, section_pages):
        pdf.add_page()
        section_pages['Profile Graph'] = pdf.page_no()
        if hasattr(pdf, 'bookmark'):
            pdf.bookmark('Profile Graph', level=0)
        pdf.set_font(font, "B", 14)
        pdf.cell(0, 10, "Profile Graph", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(font, "", 11)
        pdf.cell(0, 8, "Elevation profile generated from survey data.", new_x="LMARGIN", new_y="NEXT")
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_img:
            graph_image_path = tmp_img.name
            fig.savefig(graph_image_path, dpi=300)
        try:
            pdf.image(graph_image_path, x=10, y=pdf.get_y(), w=190)
        finally:
            if os.path.exists(graph_image_path):
                try:
                    os.remove(graph_image_path)
                except Exception as e:
                    logger.error(f"Failed to remove temporary graph image: {e}")

    def _add_pdf_calc_details(self, pdf, font, sections, section_pages):
        if sections.get('calc_details'):
            pdf.add_page()
            section_pages['Calculation Details'] = pdf.page_no()
            if hasattr(pdf, 'bookmark'):
                pdf.bookmark('Calculation Details', level=0)
            pdf.set_font(font, "B", 14)
            pdf.cell(0, 10, "Calculation Details", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font, "", 10)
            calc_method = self.master.calculation_method.get()
            pdf.multi_cell(0, 8, f"Calculation method used: {calc_method}")
            pdf.ln(5)
            if calc_method == "HI":
                pdf.set_font(font, 'B', 12)
                pdf.cell(0, 10, "Height of Instrument (HI) Method Walkthrough", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font(font, '', 10)
                pdf.multi_cell(0, 5, "1. The first RL is the starting benchmark.\n"
                                     "2. The Height of Instrument (HI) is calculated by adding the Backsight (BS) to the known RL.\n"
                                     "3. The RL of subsequent points is found by subtracting the Foresight (FS) or Intersight (IS) from the HI.\n"
                                     "4. At a change point, a new HI is calculated from the new BS reading.")
            elif calc_method == "RF":
                pdf.set_font(font, 'B', 12)
                pdf.cell(0, 10, "Rise and Fall (RF) Method Walkthrough", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font(font, '', 10)
                pdf.multi_cell(0, 5, "1. The difference between consecutive staff readings determines the Rise or Fall.\n"
                                     "2. If the previous reading is greater than the current, it's a Rise.\n")
            pdf.ln(5)

    def _add_pdf_checks(self, pdf, font, sections, section_pages):
        pdf.add_page()
        section_pages['Checks'] = pdf.page_no()
        if hasattr(pdf, 'bookmark'):
            pdf.bookmark('Checks', level=0)
        pdf.set_font(font, "B", 14)
        pdf.cell(0, 10, "Arithmetic Checks", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(font, "", 10)
        stats = self.master.calculator.stats
        if stats:
            check_ok = "OK" if not stats['arith_failed'] else "FAIL"
            pdf.multi_cell(0, 8, "The arithmetic check ensures the integrity of the leveling calculations. It is performed as follows:")
            pdf.ln(2)
            if self.master.calculation_method.get() == "HI":
                pdf.multi_cell(0, 8, f"Sum of Backsights (BS): {stats['sum_bs']:.{self.settings['precision']}f}")
                pdf.multi_cell(0, 8, f"Sum of Foresights (FS): {stats['sum_fs']:.{self.settings['precision']}f}")
                pdf.multi_cell(0, 8, f"Sum(BS) - Sum(FS): {stats['arith_check']:.{self.settings['precision']}f}")
            else:
                pdf.multi_cell(0, 8, f"Sum of Rises: {stats['sum_rise']:.{self.settings['precision']}f}")
                pdf.multi_cell(0, 8, f"Sum of Falls: {stats['sum_fall']:.{self.settings['precision']}f}")
                pdf.multi_cell(0, 8, f"Sum(Rise) - Sum(Fall): {stats['arith_check']:.{self.settings['precision']}f}")
            pdf.multi_cell(0, 8, f"Last RL - First RL: {stats['rl_diff']:.{self.settings['precision']}f}")
            pdf.ln(2)
            pdf.set_font(font, 'B', 12)
            # Set color based on check result
            if check_ok == "OK":
                pdf.set_text_color(0, 150, 0)  # Green
            else:
                pdf.set_text_color(255, 0, 0)   # Red
            pdf.multi_cell(0, 8, f"Check Result: {check_ok}")
            pdf.set_text_color(0, 0, 0)  # Reset to black
        else:
            pdf.multi_cell(0, 8, "No calculation statistics available to perform checks.")
        pdf.ln(5)

    def _fill_pdf_toc(self, pdf, font, toc_page, section_pages):
        if toc_page:
            current_page = pdf.page_no()
            pdf.page = toc_page
            pdf.set_y(pdf.get_y() + 10)
            pdf.set_font(font, "", 12)
            for section, page in section_pages.items():
                pdf.cell(0, 8, f"{section} ............................................. {page}", new_x="LMARGIN", new_y="NEXT", link=pdf.add_link())
            pdf.page = current_page

    def _apply_pdf_security(self, file_path, password, metadata, surveyor, project_name, summary):
        if not password and not metadata:
            return
        if PyPDF2 is None:
            QMessageBox.warning(self.master, "PDF Features Unavailable", "PyPDF2 is not installed. Password and metadata features are disabled.")
            return
        try:
            reader = PyPDF2.PdfReader(file_path)
            writer = PyPDF2.PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            if password:
                writer.encrypt(password)
                logger.info("PDF encrypted with a password.")
            if metadata:
                writer.add_metadata({
                    '/Author': surveyor or '',
                    '/Title': project_name or 'Leveling Survey Report',
                    '/Subject': summary or '',
                    '/Keywords': metadata
                })
                logger.info("Custom metadata added to PDF.")
            with open(file_path, 'wb') as out_f:
                writer.write(out_f)
        except Exception as e:
            logger.error(f"Failed to apply PDF security/metadata: {e}")
            QMessageBox.warning(self.master, "PDF Post-processing Error", f"Failed to apply security/metadata: {e}")

    def generate_pdf_report(self, result_table, fig, professional=False, file_path=None, **kwargs):
        logger.info("Starting PDF report generation.")
        logger.debug(f"kwargs: {kwargs}")

        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(self.master, "Generate PDF Report", "", "PDF Documents (*.pdf)")
            if not file_path:
                logger.warning("PDF generation cancelled by user.")
                return
        try:
            font = kwargs.get('font', 'Helvetica')
            if professional:
                pdf = self._create_professional_pdf(result_table, fig, **kwargs)
            else:
                pdf = self._create_simple_pdf(result_table, fig, **kwargs)
            
            pdf.output(file_path)
            self._apply_pdf_security(file_path, kwargs.get('password'), kwargs.get('metadata'), kwargs.get('surveyor'), kwargs.get('project_name'), kwargs.get('summary'))
            
            QMessageBox.information(self.master, LANG["export_success"], f"{LANG['pdf_report_success']}\n{file_path}")
            logger.info(f"Successfully generated PDF report at: {file_path}")

        except Exception as e:
            logger.error(f"Failed to generate PDF report: {e}", exc_info=True)
            QMessageBox.critical(self.master, LANG["pdf_generation_error"], f"{LANG['failed_pdf_report']}\n{e}")

    def _create_professional_pdf(self, result_table, fig, **kwargs):
        font = kwargs.get('font', 'Helvetica')
        color_scheme = kwargs.get('color_scheme', 'Default')
        summary = kwargs.get('summary', '')
        sections = kwargs.get('sections', {})
        
        pdf = PDF(project_name=kwargs.get('project_name'), surveyor=kwargs.get('surveyor'), date=kwargs.get('date'), logo_path=kwargs.get('logo_path'))
        pdf.set_auto_page_break(auto=True, margin=15)
        section_pages = {}

        self._add_pdf_cover_page(pdf, **kwargs)
        toc_page = self._add_pdf_toc(pdf, font, sections)
        self._add_pdf_summary(pdf, font, summary, section_pages)
        self._add_pdf_results_table(pdf, font, result_table, color_scheme, section_pages)
        self._add_pdf_graph(pdf, font, fig, section_pages)
        self._add_pdf_calc_details(pdf, font, sections, section_pages)
        self._add_pdf_checks(pdf, font, sections, section_pages)
        self._fill_pdf_toc(pdf, font, toc_page, section_pages)
        
        return pdf

    def _create_simple_pdf(self, result_table, fig, **kwargs):
        font = kwargs.get('font', 'Helvetica')
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font(font, "B", 16)
        pdf.cell(0, 10, "Leveling Survey Report", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font(font, "", 10)
        pdf.cell(0, 5, f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(10)
        
        section_pages = {}
        self._add_pdf_results_table(pdf, font, result_table, "Default", section_pages)
        self._add_pdf_graph(pdf, font, fig, section_pages)
        
        return pdf


    def _perform_csv_import(self, file_path, mapping, has_header, data_handler, progress_bar):
        all_rows = []
        try:
            with open(file_path, 'r', newline='') as f:
                reader = csv.reader(f)
                if has_header:
                    next(reader, None)
                for row in reader:
                    # Skip empty or malformed rows
                    if not row or all(cell.strip() == '' for cell in row):
                        continue
                    if len(row) < max(mapping.values(), default=0) + 1:
                        continue
                    all_rows.append(row)
            data_handler(all_rows, mapping)
        except Exception as e:
            QMessageBox.critical(self.master, LANG["import_error"], f"{LANG['failed_import_csv']}\n{e}")

    def _update_recent_files(self, file_path):
        # Implement your logic to update recent files in settings
        pass 
